import os
import time
import shutil
import pandas as pd
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.common.exceptions import (
    StaleElementReferenceException,
    ElementNotInteractableException,
    InvalidElementStateException,
    TimeoutException,
    NoSuchElementException,
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

driver_path = r"edgedriver\msedgedriver.exe"
excel_path = "financial_analytics.xlsx"
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
project_folder = os.getcwd()
reports_folder = os.path.join(project_folder, "All_reports")
os.makedirs(reports_folder, exist_ok=True)

def unique_filename(folder, filename):
    """
    If `filename` exists in `folder`, append a counter *without* any separator:
    base -> base1 -> base2 -> ...
    Example: code_08-2025_TB.xlsx -> code_08-2025_TB1.xlsx -> code_08-2025_TB2.xlsx
    """
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(folder, filename)
    if not os.path.exists(candidate):
        return candidate
    n = 1
    while True:
        candidate = os.path.join(folder, f"{base}{n}{ext}")
        if not os.path.exists(candidate):
            return candidate
        n += 1
    return candidate

df = pd.read_excel(excel_path)
required_cols = {"Codes", "Report_type", "From_period", "To_period"}
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required columns in Excel: {missing}")

df["FromFormatted"] = pd.to_datetime(df["From_period"]).dt.strftime("%m/%Y")
df["ToFormatted"]   = pd.to_datetime(df["To_period"]).dt.strftime("%m/%Y")

options = Options()
options.use_chromium = True
options.add_argument("--start-maximized")
options.add_argument("--log-level=3")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
wait = WebDriverWait(driver, 25)

def reenter_target_iframe():
    """Make sure we’re inside the latest report iframe (the page often replaces it)."""
    driver.switch_to.default_content()
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    if not frames:
        raise RuntimeError("No iframe found on page.")
    driver.switch_to.frame(frames[-1])

def js_set_value(el, value):
    """Set input value via JS and dispatch events."""
    driver.execute_script(
        """
        const el = arguments[0], val = arguments[1];
        el.removeAttribute('readonly');
        el.removeAttribute('disabled');
        el.value = val;
        el.dispatchEvent(new Event('input', {bubbles:true}));
        el.dispatchEvent(new Event('change', {bubbles:true}));
        """,
        el, value
    )

def safe_type(by, locator, value, click_first=True, use_js_fallback=True, clear_first=True):
    """
    Robust setter for inputs: waits visible+enabled, scrolls, tries clear+send_keys, then JS fallback.
    Returns True if it believes value is set.
    """
    try:
        el = wait.until(EC.presence_of_element_located((by, locator)))
        wait.until(EC.visibility_of_element_located((by, locator)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        if click_first:
            try:
                wait.until(EC.element_to_be_clickable((by, locator)))
                el.click()
            except Exception:
                pass
        if clear_first:
            try:
                el.clear()
            except (InvalidElementStateException, ElementNotInteractableException):
                driver.execute_script("arguments[0].value='';", el)
        try:
            el.send_keys(value)
            return True
        except (InvalidElementStateException, ElementNotInteractableException):
            if use_js_fallback:
                js_set_value(el, value)
                return True
            return False
    except (TimeoutException, StaleElementReferenceException, NoSuchElementException):
        return False

def wait_for_new_xlsx(before_set, timeout=60, stable_wait=2):
    """
    Wait for a new .xlsx to appear in Downloads.
    Ensures the file size is stable for stable_wait seconds before returning path.
    """
    end = time.time() + timeout
    while time.time() < end:
        after = set(os.listdir(downloads_folder))
        new_files = [f for f in (after - before_set) if f.lower().endswith(".xlsx")]
        if new_files:
            # most recent among new ones
            paths = [os.path.join(downloads_folder, f) for f in new_files]
            candidate = max(paths, key=os.path.getctime)
            size1 = os.path.getsize(candidate)
            time.sleep(stable_wait)
            size2 = os.path.getsize(candidate)
            if size1 == size2:
                return candidate
        time.sleep(1)
    return None

def ensure_checkbox_checked(checkbox_id="SupressZero_CheckBox"):
    """Ensure the target checkbox is checked. Safe if it's already checked."""
    try:
        cb = wait.until(EC.presence_of_element_located((By.ID, checkbox_id)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)

        def _is_checked(elem):
            try:
                return elem.is_selected() or (str(elem.get_attribute("checked")).lower() in ("true", "checked"))
            except Exception:
                return False

        if not _is_checked(cb):
            # Try normal click
            try:
                wait.until(EC.element_to_be_clickable((By.ID, checkbox_id))).click()
            except Exception:
                # Fallback: click the label if present, else JS click
                try:
                    lbl = driver.find_element(By.XPATH, f"//label[@for='{checkbox_id}']")
                    driver.execute_script("arguments[0].click();", lbl)
                except Exception:
                    driver.execute_script("arguments[0].click();", cb)


    except Exception as e:
        # If the checkbox isn't present for some report types/screens, just log and continue
        print(f"ℹ️ Checkbox '{checkbox_id}' not found or not clickable now: {e}")

wb = load_workbook(excel_path)
ws = wb.active
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
had_failures = False

TREE_BY_TYPE = {
    "Trial Balance":                 "ysi_tb",
    "Balance Sheet":                 "ysi_bs",
    "Income Statement":              "ysi_is",
    "Budget Comparison (with PTD)":  "ysi_is",
    "12 Month Statement":            "ysi_is",   
}
SUFFIX_BY_TYPE = {
    "Trial Balance":                 "TB",
    "Balance Sheet":                 "BS",
    "Income Statement":              "IS",
    "Budget Comparison (with PTD)":  "BC_PTD",
    "12 Month Statement":            "MS12",               
}

driver.get("https://www.yardiasp14.com/66553dolphin/pages/menu.aspx")
input("🔐 Please log in manually and press ENTER here to continue...")

# Menu path for this script (keep as per your page)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi1"]/a'))).click()
from selenium.webdriver import ActionChains
ActionChains(driver).move_to_element(
    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi1-10"]/a')))
).perform()
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm1-10"]/li[5]/a'))).click()
time.sleep(2)
reenter_target_iframe()

for idx, row in df.iterrows():
    code = str(row["Codes"]).strip()
    report_type = str(row["Report_type"]).strip()
    from_str = (str(row["FromFormatted"]).strip()
                if pd.notna(row["FromFormatted"]) else "")
    to_str   = (str(row["ToFormatted"]).strip()
                if pd.notna(row["ToFormatted"]) else "")

    tree_id = TREE_BY_TYPE.get(report_type, "").strip()
    suffix  = SUFFIX_BY_TYPE.get(report_type, "REP").strip()

    print(f"\n➡️ Processing: {code} | Report: {report_type} | Period: {from_str} → {to_str}")

    success = False
    for attempt in range(1, 4):
        try:
            print(f"   🔁 Attempt {attempt}/3")

            # The iframe may refresh after each report; always re-enter
            reenter_target_iframe()
            # Make sure the "Suppress Zero" checkbox is checked BEFORE filling inputs
            ensure_checkbox_checked("SupressZero_CheckBox")


            # 1) Property code
            if not safe_type(By.ID, "PropertyID_LookupCode", code):
                raise RuntimeError("Could not set PropertyID_LookupCode")

            # 2) Report type FIRST (some types toggle date fields)
            try:
                ddl = wait.until(EC.presence_of_element_located((By.ID, "ReportNum_DropDownList")))
                Select(ddl).select_by_visible_text(report_type)
                # Some report type changes re-render the form; ensure the checkbox stayed checked
                ensure_checkbox_checked("SupressZero_CheckBox")

            except Exception:
                raise RuntimeError("Could not select ReportNum_DropDownList")

            # 3) Book = Accrual
            if not safe_type(By.ID, "BookID_LookupCode", "Accrual"):
                raise RuntimeError("Could not set BookID_LookupCode")

            # 4) TreeID depends on report type
            if tree_id:
                if not safe_type(By.ID, "TreeID_LookupCode", tree_id):
                    raise RuntimeError("Could not set TreeID_LookupCode")

            if report_type != "Balance Sheet" and from_str:
                if not safe_type(By.ID, "FromMMYY_TextBox", from_str):
                    raise RuntimeError("Could not set FromMMYY_TextBox")
            if to_str:
                if not safe_type(By.ID, "ToMMYY_TextBox", to_str):
                    raise RuntimeError("Could not set ToMMYY_TextBox")


            # 6) Display then Excel
            disp = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Display_Button"]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", disp)
            disp.click()
            time.sleep(2)

            before = set(os.listdir(downloads_folder))
            excel_btn = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Excel_Button"]')))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", excel_btn)
            excel_btn.click()

            downloaded = wait_for_new_xlsx(before_set=before, timeout=60, stable_wait=2)
            if downloaded:
                # Use From if present, otherwise fall back to To
                name_period = (to_str).replace("/", "-") if (to_str) else "NA"
                new_name = f"{code}_{name_period}_{suffix}.xlsx"
                unique_path = unique_filename(reports_folder, new_name)  # <<< uses _1, _2, ...
                shutil.move(downloaded, unique_path)
                print(f"   ✅ Saved: {os.path.basename(unique_path)}")
                success = True
                break
            else:
                print("   ⚠️ No new .xlsx detected.")

        except Exception as e:
            print(f"   ❌ Error: {e}")
            time.sleep(2)

    if not success:
        print(f"   ❌ Failed after 3 attempts: {code}")
        had_failures = True
        # Highlight Excel row
        excel_row_index = idx + 2  # header offset
        for cell in ws[excel_row_index]:
            cell.fill = red_fill

# Save Excel only if failures occurred
if had_failures:
    try:
        wb.save(excel_path)
        print(f"\n📘 Excel updated (failed rows highlighted): {excel_path}")
    except PermissionError:
        print(f"\n⛔ Cannot save '{excel_path}'. Please close the file if it's open and run again.")
wb.close()
print("Report downloads finished. You can exit this command window.")
driver.quit()
