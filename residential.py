import os
import time
import shutil
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Setup Paths ===
driver_path = r"edgedriver\msedgedriver.exe"
excel_path = "residential.xlsx"
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
project_folder = os.getcwd()
reports_folder = os.path.join(project_folder, "All_reports")
os.makedirs(reports_folder, exist_ok=True)

# === Read Excel ===
df = pd.read_excel(excel_path)
df["FromFormatted"] = pd.to_datetime(df["Date"]).dt.strftime("%m/%d/%Y")
df["ToFormatted"] = pd.to_datetime(df["Month"]).dt.strftime("%m/%Y")

# === Setup Edge ===
options = Options()
options.use_chromium = True
options.add_argument("--start-maximized")
options.add_argument("--log-level=3")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
wait = WebDriverWait(driver, 20)

# === Login ===
driver.get("https://www.yardiasp14.com/66553dolphin/pages/menu.aspx")
input("🔐 Please log in manually and press ENTER here to continue...")

# === Navigation to Report Page ===
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi1"]/a'))).click()
actions = webdriver.ActionChains(driver)
actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi1-2"]/a')))).perform()
time.sleep(1)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm1-2"]/li[2]/a'))).click()
time.sleep(2)
driver.switch_to.frame(driver.find_elements(By.TAG_NAME, "iframe")[-1])

# === Excel Setup for Failed Rows ===
wb = load_workbook(excel_path)
ws = wb.active
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# === Main Processing Loop ===
def get_latest_download(folder):
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".xlsx")]
    if not files:
        return None
    return max(files, key=os.path.getctime)

for index, row in df.iterrows():
    code = row["Codes"]
    from_period = row["FromFormatted"]
    to_period = row["ToFormatted"]

    print(f"\n📄 Processing Property: '{code}' | Period: {from_period} to {to_period}")
    success = False

    for attempt in range(3):
        try:
            print(f"➡️ Attempt {attempt + 1}...")

            # Fill Inputs
            driver.find_element(By.ID, "PropLookup_LookupCode").clear()
            driver.find_element(By.ID, "PropLookup_LookupCode").send_keys(code)
            driver.find_element(By.ID, "Date2_TextBox").clear()
            driver.find_element(By.ID, "Date2_TextBox").send_keys(from_period)
            driver.find_element(By.ID, "MMYY2_TextBox").clear()
            driver.find_element(By.ID, "MMYY2_TextBox").send_keys(to_period)
            Select(driver.find_element(By.ID, "ReportType_DropDownList")).select_by_visible_text("Rent Roll with Lease Charges")
            Select(driver.find_element(By.ID, "SummarizeBy_DropDownList")).select_by_visible_text("Unit")

            # Click Display
            wait.until(EC.element_to_be_clickable((By.ID, "Display_Button"))).click()
            print("📊 Display clicked")
            time.sleep(2)
            # Click Excel and wait for download
            before = set(os.listdir(downloads_folder))
            wait.until(EC.element_to_be_clickable((By.ID, "Excel_Button"))).click()
            print("⬇️ Download initiated...")

            timeout = time.time() + 30
            downloaded_file = None
            while time.time() < timeout:
                after = set(os.listdir(downloads_folder))
                new_files = after - before
                if new_files:
                    for f in new_files:
                        if f.endswith(".xlsx"):
                            downloaded_file = os.path.join(downloads_folder, f)
                            break
                if downloaded_file:
                    break
                time.sleep(1)

            if downloaded_file:
                new_name = f"{code}_{from_period.replace('/', '-')}_PR.xlsx"
                shutil.move(downloaded_file, os.path.join(reports_folder, new_name))
                print(f"✅ Saved as: {new_name}")
                success = True
                break
            else:
                print("❌ Download not detected.")

        except Exception as e:
            print(f"⚠️ Error during attempt {attempt + 1}: {e}")
            time.sleep(2)

    if not success:
        print(f"❌ All attempts failed for property: {code}")
        for cell in ws[index + 2]:
            cell.fill = fill_red

# === Save Excel with failed rows ===
try:
    wb.save(excel_path)
    print(f"\n📘 Excel updated with any failed rows highlighted: {excel_path}")
except PermissionError:
    print(f"\n⛔ Cannot save '{excel_path}'. Please close the file if it's open and try again.")

wb.close()
print("Report downloads finished. You can exit this command window.")
driver.quit()
