import os
import time
import shutil
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Setup Paths ===
driver_path = r"edgedriver\msedgedriver.exe"
excel_path = "affordable_report.xlsx"
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
project_folder = os.getcwd()
reports_folder = os.path.join(project_folder, "All_reports")
os.makedirs(reports_folder, exist_ok=True)

# === Setup Edge ===
options = Options()
options.use_chromium = True
options.add_argument("--start-maximized")
options.add_argument("--log-level=3")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

# === Initialize Driver ===
driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
wait = WebDriverWait(driver, 30)
actions = ActionChains(driver)

# === Load Excel ===
df = pd.read_excel(excel_path)
wb = load_workbook(excel_path)
ws = wb.active
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

# === Helper Functions ===
def get_latest_download(folder):
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    return max(files, key=os.path.getctime) if files else None

def unique_filename(folder, filename):
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(folder, filename)
    n = 1
    while os.path.exists(candidate):
        candidate = os.path.join(folder, f"{base}({n}){ext}")
        n += 1
    return candidate

# Wait until a *different/newer* xlsx becomes the latest after we click
def wait_new_latest_xlsx(folder, prev_path, prev_mtime, timeout=12, stable_wait=1.0):
    end = time.time() + timeout
    while time.time() < end:
        path = get_latest_download(folder)
        if path:
            try:
                mtime = os.path.getmtime(path)
                if (prev_path is None) or (path != prev_path) or (mtime > (prev_mtime + 0.25)):
                    s1 = os.path.getsize(path)
                    time.sleep(stable_wait)
                    s2 = os.path.getsize(path)
                    if s1 == s2:
                        return path
            except FileNotFoundError:
                pass
        time.sleep(0.25)
    return None

# NEW: wait until the View Report link is truly ready (visible, enabled, and with a real href)
def wait_view_report_ready(max_wait=45):
    xpath = '//*[@id="TableWriter1_Row0"]/td/a'
    end = time.time() + max_wait
    while time.time() < end:
        try:
            el = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            if el.is_displayed() and el.is_enabled():
                href = el.get_attribute("href") or ""
                # your popup URL shows SysShuttleDisplayHandler.ashx?FileName=...
                if "SysShuttleDisplayHandler" in href or "FileName=" in href or href.startswith("http"):
                    return el
        except (TimeoutException, StaleElementReferenceException):
            pass
        time.sleep(0.3)
    raise TimeoutException("View Report link not ready in time.")

# === Login Flow ===
driver.get("https://www.yardiasp14.com/66553dolphin/pages/menu.aspx")
input("🔐 Please log in manually and press ENTER here to continue...")

wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi0"]')))).perform()
time.sleep(0.8)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
time.sleep(2)
driver.switch_to.default_content()
iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
driver.switch_to.frame(iframes[0])
time.sleep(0.8)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_ctl310"]'))).click()
time.sleep(1.5)
driver.switch_to.default_content()
iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
driver.switch_to.frame(iframes[-1])

# Select report type
Select(driver.find_element(By.ID, "YsiMergeReport_DropDownList")).select_by_visible_text(
    "Affordable Rent Roll with Lease Charges (AffRntRollLsChgs)")
time.sleep(4)
driver.switch_to.default_content()
iframe_after_selection = driver.find_elements(By.TAG_NAME, "iframe")[-1]
driver.switch_to.frame(iframe_after_selection)

# main input window (where we enter inputs)
main_window = driver.current_window_handle

# === Main Loop ===
for index, row in df.iterrows():
    prop_code = str(row.get("Codes", "")).strip()
    date = row.get("Date", "")
    month = row.get("Month", "")

    print(f"\n📄 Processing: ({prop_code})")

    # For renaming: month string as MM-YYYY
    try:
        month_for_name = pd.to_datetime(month).strftime("%m-%Y")
    except Exception:
        month_for_name = "NA"

    success = False
    for attempt in range(3):
        try:
            print(f"➡️ Attempt {attempt + 1}")

            driver.switch_to.default_content()
            iframe = driver.find_elements(By.TAG_NAME, "iframe")[-1]
            driver.switch_to.frame(iframe)

            driver.find_element(By.ID, "Ysi4114_LookupCode").clear()
            driver.find_element(By.ID, "Ysi4114_LookupCode").send_keys(prop_code)
            time.sleep(1)
            driver.find_element(By.ID, "Ysi4117_TextBox").clear()
            driver.find_element(By.ID, "Ysi4117_TextBox").send_keys(pd.to_datetime(date).strftime("%m/%d/%Y"))
            time.sleep(1)
            driver.find_element(By.ID, "Ysi4118_TextBox").clear()
            driver.find_element(By.ID, "Ysi4118_TextBox").send_keys(pd.to_datetime(month).strftime("%m/%Y"))
            time.sleep(1)

            Select(driver.find_element(By.ID, "YsiMergeReport_DropDownList")).select_by_visible_text(
                "Affordable Rent Roll with Lease Charges (AffRntRollLsChgs)")
            time.sleep(1)
            Select(driver.find_element(By.ID, "Ysi4122_DropDownList")).select_by_visible_text("Unit")
            time.sleep(1)
            Select(driver.find_element(By.ID, "YsiOutpuType_DropDownList")).select_by_visible_text("Excel")
            time.sleep(1)

            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSubmit_Button"]'))).click()
            print("⏳ Waiting for report to process...")

            # ---- NEW robust wait for the real View Report link ----
            view_button = wait_view_report_ready(max_wait=45)

            # Snapshot current latest xlsx BEFORE clicking
            prev_latest = get_latest_download(downloads_folder)
            prev_mtime = os.path.getmtime(prev_latest) if prev_latest and os.path.exists(prev_latest) else 0

            handles_before = driver.window_handles
            view_button.click()
            print("📅 View Report clicked.")

            # Optional: detect popup and switch so Edge initiates the download there
            popup_handle = None
            try:
                WebDriverWait(driver, 5).until(EC.new_window_is_opened(handles_before))
                popup_candidates = [h for h in driver.window_handles if h not in handles_before]
                if popup_candidates:
                    popup_handle = popup_candidates[0]
            except Exception:
                popup_handle = None
            if popup_handle:
                try:
                    driver.switch_to.window(popup_handle)
                except Exception:
                    popup_handle = None

            # Wait for a different/newer xlsx to appear and finish
            downloaded = wait_new_latest_xlsx(
                downloads_folder, prev_path=prev_latest, prev_mtime=prev_mtime, timeout=12, stable_wait=1.0
            )

            # Close popups and return to main
            for h in list(driver.window_handles):
                if h != main_window:
                    try:
                        driver.switch_to.window(h)
                        driver.close()
                    except Exception:
                        pass
            driver.switch_to.window(main_window)

            # Rename+move if found
            if downloaded:
                new_name = f"{prop_code}_{month_for_name}_AR.xlsx"  # codes_MM-YYYY_AR
                dest_path = unique_filename(reports_folder, new_name)
                shutil.move(downloaded, dest_path)
                print(f"✅ Saved as: {os.path.basename(dest_path)}")
                success = True
                break
            else:
                print("⚠️ No new .xlsx detected after View Report.")

        except Exception as e:
            print(f"⚠️ Attempt {attempt + 1} failed: {e}")
            try:
                driver.switch_to.window(main_window)
                iframe = driver.find_elements(By.TAG_NAME, "iframe")[-1]
                driver.switch_to.frame(iframe)
                print("↩️ Attempting to return to main window after failure...")
            except:
                pass
            time.sleep(2)

    if not success:
        print(f"❌ All 3 attempts failed for: {prop_code}")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=index + 2, column=col).fill = red_fill

# === Save Excel ===
if any(cell.fill == red_fill for row in ws.iter_rows(min_row=2) for cell in row):
    try:
        wb.save(excel_path)
        print(f"\n📘 Excel updated with failed rows highlighted: {excel_path}")
    except PermissionError:
        print(f"\n⛔ Cannot save '{excel_path}'. Please close the file if it's open and try again.")
else:
    print("\n✅ All reports downloaded successfully. No highlights needed.")

wb.close()
print("Report downloads finished. You can exit this command window.")
driver.quit()