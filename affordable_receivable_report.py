import os
import time
import shutil
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Setup Paths ===
driver_path = r"edgedriver\msedgedriver.exe"
excel_path = "affordable_receivable_report.xlsx"
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
project_folder = os.getcwd()
reports_folder = os.path.join(project_folder, "All_reports")
os.makedirs(reports_folder, exist_ok=True)

# === Read Excel ===
df = pd.read_excel(excel_path)

# Handle month column dynamically
MONTH_CANDIDATES = ["Month", "From_period", "From", "Period", "MMYY", "As_of_Month"]
month_col = next((c for c in MONTH_CANDIDATES if c in df.columns), None)
if not month_col:
    raise ValueError(f"No month-like column found. Got columns: {list(df.columns)}")

# Format as mm/YYYY
df["FromFormatted"] = pd.to_datetime(df[month_col], errors="coerce").dt.strftime("%m/%Y")

# === Setup Edge ===
options = Options()
options.use_chromium = True
options.add_argument("--start-maximized")
options.add_argument("--log-level=3")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
wait = WebDriverWait(driver, 20)
actions = ActionChains(driver)

# === Login ===
driver.get("https://www.yardiasp14.com/66553dolphin/pages/menu.aspx")
input("🔐 Please log in manually and press ENTER here to continue...")

wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mi0"]/a'))).click()
actions.move_to_element(wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mi0"]')))).perform()
time.sleep(0.8)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sm0"]/li[5]/a'))).click()
time.sleep(2)
driver.switch_to.default_content()
iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
driver.switch_to.frame(iframes[0])
time.sleep(0.8)
wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="_ctl313"]'))).click()
time.sleep(1.5)
driver.switch_to.default_content()
iframes = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
driver.switch_to.frame(iframes[-1])

# === Excel Setup for Failed Rows ===
wb = load_workbook(excel_path)
ws = wb.active
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# === Helpers ===
def unique_filename(folder, filename):
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(folder, filename)
    n = 1
    while os.path.exists(candidate):
        candidate = os.path.join(folder, f"{base}({n}){ext}")
        n += 1
    return candidate

def wait_for_new_xlsx(before_set, timeout=30, stable_wait=1):
    """
    Wait for a new .xlsx to appear in Downloads after clicking 'Excel'.
    Ensures size is stable for `stable_wait` seconds before returning.
    """
    end = time.time() + timeout
    while time.time() < end:
        after = set(os.listdir(downloads_folder))
        new_files = [f for f in (after - before_set) if f.lower().endswith(".xlsx")]
        if new_files:
            # choose the newest among new files
            paths = [os.path.join(downloads_folder, f) for f in new_files]
            candidate = max(paths, key=os.path.getctime)
            try:
                s1 = os.path.getsize(candidate)
                time.sleep(stable_wait)
                s2 = os.path.getsize(candidate)
                if s1 == s2:
                    return candidate
            except FileNotFoundError:
                pass
        time.sleep(0.5)
    return None

def run_once_for_subsidy(code, period_str, subsidy_text, suffix_tag):
    """
    Select HUD subsidy option, click Display, click Excel,
    wait for the new file, and move+rename.
    Returns True on success, False otherwise.
    """
    # Set HUD Subsidies value
    Select(driver.find_element(By.ID, "cmbHUDSubsidies_DropDownList")).select_by_visible_text(subsidy_text)
    time.sleep(0.4)

    # Display
    wait.until(EC.element_to_be_clickable((By.ID, "Display_Button"))).click()
    print(f"📊 Display clicked ({subsidy_text})")
    time.sleep(2)

    # Excel → detect new file
    before = set(os.listdir(downloads_folder))
    wait.until(EC.element_to_be_clickable((By.ID, "Excel_Button"))).click()
    print(f"⬇️ Download initiated ({subsidy_text})...")

    downloaded_file = wait_for_new_xlsx(before_set=before, timeout=30, stable_wait=1)
    if downloaded_file:
        new_name = f"{code}_{period_str.replace('/', '-')}_ARR_{suffix_tag}.xlsx"
        dest_path = unique_filename(reports_folder, new_name)
        shutil.move(downloaded_file, dest_path)
        print(f"✅ Saved as: {os.path.basename(dest_path)}")
        return True
    else:
        print(f"❌ Download not detected for ({subsidy_text}).")
        return False

# === Main Processing Loop ===
for index, row in df.iterrows():
    code = str(row["Codes"]).strip()
    period = row["FromFormatted"]

    if pd.isna(period) or not str(period).strip():
        print(f"⚠️ Skipping property {code}: invalid period")
        continue

    print(f"\n📄 Processing Property: '{code}' | Period: '{period}'")
    row_success = True  # will be set False if any of the two runs fails

    for attempt in range(3):
        try:
            print(f"➡️ Attempt {attempt + 1}...")

            # Fill Inputs (common for both runs)
            driver.find_element(By.ID, "PropLookup_LookupCode").clear()
            driver.find_element(By.ID, "PropLookup_LookupCode").send_keys(code)

            Select(driver.find_element(By.ID, "ReportType_DropDownList")).select_by_visible_text("Receivable Aging Summary")
            Select(driver.find_element(By.ID, "SummarizeBy_DropDownList")).select_by_visible_text("Resident")

            driver.find_element(By.ID, "MMYY2_TextBox").clear()
            driver.find_element(By.ID, "MMYY2_TextBox").send_keys(period)

            # === Run 1: HUD Subsidies = Include → ..._ARR_I.xlsx
            ok_include = run_once_for_subsidy(code, period, "Include", "I")

            # === Run 2: HUD Subsidies = Exclude → ..._ARR_E.xlsx
            ok_exclude = run_once_for_subsidy(code, period, "Exclude", "E")

            if ok_include and ok_exclude:
                # Both succeeded for this row
                break
            else:
                row_success = False
                print("⚠️ One of the two downloads failed; retrying the whole row...")
                time.sleep(2)

        except Exception as e:
            row_success = False
            print(f"⚠️ Error during attempt {attempt + 1}: {e}")
            time.sleep(2)

    if not row_success:
        print(f"❌ Final status: at least one download failed for property: {code}")
        for cell in ws[index + 2]:
            cell.fill = fill_red

# === Save Excel with failed rows ===
try:
    wb.save(excel_path)
    print(f"\n📘 Excel updated with any failed rows highlighted: {excel_path}")
except PermissionError:
    print(f"\n⛔ Cannot save '{excel_path}'. Please close the file if it's open and try again.")

wb.close()
print("Report downloads finished. You can exit this command window.")
driver.quit()
